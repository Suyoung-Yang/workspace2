from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "nadosheet"
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # A1 셀 정보 출력
print(ws["A1"].value)
print(ws["A10"].value) # 값이 없을 때는 none 출력

print(ws.cell(row = 1, column = 1).value) # a1셀과 같음
print(ws.cell(row = 1, column = 2).value) # b1셀과 같음

c = ws.cell(column=3,row=1, value=10) # c3에 10 입력. ws["c1"] = 10
print(c.value)


from random import *

# 반복문 이용해서 랜덤 숫자 넣기
index = 1
for x in range(1,11):
    for y in range(1,11):
        # ws.cell(row=x,column=y,value=randint(0,100)) # 0 ~ 100 사이 숫자
        ws.cell(row=x,column=y,value = index)
        index +=1





wb.save("sample.xlsx")
